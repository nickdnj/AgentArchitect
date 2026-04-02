#!/usr/bin/env python3
"""
Wharfside Manor Budget Analysis — 2025 vs 2026 Year-over-Year Comparison
Combines two AppFolio budget exports into one analysis workbook with charts.
"""

import openpyxl
from openpyxl.styles import Font, Alignment, numbers, PatternFill, Border, Side
from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter
from copy import copy

# --- Configuration ---
FILE_2025 = "/Users/nickd/Downloads/annual_budget_comparative-20251201.xlsx"
FILE_2026 = "/Users/nickd/Downloads/annual_budget_comparative-20260401.xlsx"
OUTPUT = "/Users/nickd/Downloads/Wharfside_Budget_Analysis_2025_2026.xlsx"

HEADER_ROW = 11  # Row with column headers in both source files
ACCT_FMT = '#,##0.00_);(#,##0.00)'
PCT_FMT = '0.00%'

# Months of data in the 2026 file (Jan-Apr = 4 months)
MONTHS_2026 = 4

# --- Helpers ---

def read_source(path):
    """Read an AppFolio export and return (worksheet_object, workbook_object)."""
    wb = openpyxl.load_workbook(path)
    return wb, wb.active


def copy_sheet(src_ws, dst_ws):
    """Copy all cells (values + formatting) from source to destination worksheet."""
    for row in src_ws.iter_rows(min_row=1, max_row=src_ws.max_row, max_col=src_ws.max_column):
        for cell in row:
            new_cell = dst_ws.cell(row=cell.row, column=cell.column, value=cell.value)
            if cell.has_style:
                new_cell.font = copy(cell.font)
                new_cell.border = copy(cell.border)
                new_cell.fill = copy(cell.fill)
                new_cell.number_format = cell.number_format
                new_cell.protection = copy(cell.protection)
                new_cell.alignment = copy(cell.alignment)
    # Copy merged cells
    for mc in src_ws.merged_cells.ranges:
        dst_ws.merge_cells(str(mc))
    # Copy column widths
    for col_letter, dim in src_ws.column_dimensions.items():
        dst_ws.column_dimensions[col_letter].width = dim.width


def extract_data_rows(ws):
    """
    Parse the AppFolio export into a list of dicts.
    Returns rows from HEADER_ROW+1 to the end, skipping blank/section-header rows.
    Each dict: {account_name, ytd_actual, ytd_budget, ytd_var_dollar, ytd_var_pct, annual_budget, is_total, section}
    """
    rows = []
    current_section = None
    for row in ws.iter_rows(min_row=HEADER_ROW + 1, max_row=ws.max_row, max_col=10):
        name = row[0].value
        if name is None:
            continue
        name_stripped = name.strip()

        # Track section (Income / Expense / Other Expense)
        if name_stripped in ("Income", "Expense", "Other Expense"):
            current_section = name_stripped
            continue

        # Get YTD and Annual columns
        ytd_actual = row[5].value
        ytd_budget = row[6].value
        ytd_var_dollar = row[7].value
        ytd_var_pct = row[8].value
        annual_budget = row[9].value

        # Skip rows with no numeric data (sub-headers like "    Pools", "    Insurance")
        if ytd_actual is None and ytd_budget is None and annual_budget is None:
            continue

        is_total = name_stripped.startswith("Total ") or name_stripped.startswith("NOI") or name_stripped == "Net Income"

        rows.append({
            "account_name": name_stripped,
            "ytd_actual": ytd_actual if ytd_actual else 0.0,
            "ytd_budget": ytd_budget if ytd_budget else 0.0,
            "ytd_var_dollar": ytd_var_dollar if ytd_var_dollar else 0.0,
            "ytd_var_pct": ytd_var_pct if ytd_var_pct else "0.00%",
            "annual_budget": annual_budget if annual_budget else 0.0,
            "is_total": is_total,
            "section": current_section,
        })
    return rows


def build_account_map(rows):
    """Build a dict keyed by account_name for quick lookup."""
    return {r["account_name"]: r for r in rows}


def set_col_widths(ws, widths):
    """Set column widths from a dict {col_index: width}."""
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w


def fmt_acct(cell):
    """Apply accounting number format."""
    cell.number_format = ACCT_FMT
    cell.alignment = Alignment(horizontal="right")


def fmt_pct(cell):
    """Apply percentage format."""
    cell.number_format = '0.00%'
    cell.alignment = Alignment(horizontal="right")


def style_header(ws, row, max_col):
    """Bold + fill the header row."""
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    thin_border = Border(
        bottom=Side(style="thin", color="1F4E79")
    )
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", wrap_text=True)


def style_total_row(ws, row, max_col):
    """Bold + top border for total rows."""
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = Font(bold=True)
        cell.border = Border(top=Side(style="thin"))


# --- Main ---

def main():
    # Load source files
    wb_2025, ws_2025 = read_source(FILE_2025)
    wb_2026, ws_2026 = read_source(FILE_2026)

    # Create output workbook
    out = openpyxl.Workbook()

    # ========== Sheet 1: 2025 Actuals ==========
    ws1 = out.active
    ws1.title = "2025 Actuals"
    copy_sheet(ws_2025, ws1)

    # ========== Sheet 2: 2026 YTD ==========
    ws2 = out.create_sheet("2026 YTD")
    copy_sheet(ws_2026, ws2)

    # ========== Sheet 3: Year-over-Year Comparison ==========
    ws3 = out.create_sheet("Year-over-Year Comparison")

    data_2025 = extract_data_rows(ws_2025)
    data_2026 = extract_data_rows(ws_2026)
    map_2025 = build_account_map(data_2025)
    map_2026 = build_account_map(data_2026)

    # Determine the union of accounts (preserving order from 2025 first, then new 2026 accounts)
    all_accounts = []
    seen = set()
    for r in data_2025:
        all_accounts.append(r["account_name"])
        seen.add(r["account_name"])
    for r in data_2026:
        if r["account_name"] not in seen:
            all_accounts.append(r["account_name"])
            seen.add(r["account_name"])

    # Title
    ws3.cell(row=1, column=1, value="Wharfside Manor — Year-over-Year Budget Comparison")
    ws3.cell(row=1, column=1).font = Font(bold=True, size=14, color="1F4E79")
    ws3.cell(row=2, column=1, value="2025 Full Year (Dec 2025) vs 2026 YTD (Apr 2026)")
    ws3.cell(row=2, column=1).font = Font(italic=True, size=11, color="555555")

    # Headers (row 4)
    headers = [
        "Account Name",
        "2025 YTD Actual",
        "2025 Annual Budget",
        "2025 Var ($)",
        "2025 Var (%)",
        "2026 YTD Actual",
        "2026 Annual Budget",
        "2026 Var ($)",
        "2026 Var (%)",
        "2026 Annualized\nProjection",
        "YoY Change ($)\n(Projected)",
        "YoY Change (%)",
    ]
    hdr_row = 4
    for ci, h in enumerate(headers, 1):
        ws3.cell(row=hdr_row, column=ci, value=h)
    style_header(ws3, hdr_row, len(headers))

    # Data rows
    r = hdr_row + 1
    # Track rows for chart data (expense line items only, non-total)
    expense_chart_rows = []  # (row_number, account_name)
    income_total_row = None
    expense_total_row = None

    for acct in all_accounts:
        d25 = map_2025.get(acct)
        d26 = map_2026.get(acct)

        ytd_25 = d25["ytd_actual"] if d25 else 0.0
        bud_25 = d25["annual_budget"] if d25 else 0.0
        var_25 = ytd_25 - bud_25
        var_pct_25 = (var_25 / bud_25) if bud_25 != 0 else 0.0

        ytd_26 = d26["ytd_actual"] if d26 else 0.0
        bud_26 = d26["annual_budget"] if d26 else 0.0
        var_26 = ytd_26 - (bud_26 * MONTHS_2026 / 12)
        var_pct_26 = (var_26 / (bud_26 * MONTHS_2026 / 12)) if bud_26 != 0 else 0.0

        # Annualized projection: YTD * 12 / months
        annualized_26 = ytd_26 * 12 / MONTHS_2026 if ytd_26 != 0 else 0.0

        # YoY change: annualized 2026 vs 2025 actual
        yoy_dollar = annualized_26 - ytd_25
        yoy_pct = (yoy_dollar / ytd_25) if ytd_25 != 0 else 0.0

        is_total = (d25 and d25["is_total"]) or (d26 and d26["is_total"])
        section = (d25 or d26 or {}).get("section", "")

        ws3.cell(row=r, column=1, value=acct)
        ws3.cell(row=r, column=2, value=ytd_25)
        ws3.cell(row=r, column=3, value=bud_25)
        ws3.cell(row=r, column=4, value=var_25)
        ws3.cell(row=r, column=5, value=var_pct_25)
        ws3.cell(row=r, column=6, value=ytd_26)
        ws3.cell(row=r, column=7, value=bud_26)
        ws3.cell(row=r, column=8, value=var_26)
        ws3.cell(row=r, column=9, value=var_pct_26)
        ws3.cell(row=r, column=10, value=annualized_26)
        ws3.cell(row=r, column=11, value=yoy_dollar)
        ws3.cell(row=r, column=12, value=yoy_pct)

        # Format currency columns
        for ci in [2, 3, 4, 6, 7, 8, 10, 11]:
            fmt_acct(ws3.cell(row=r, column=ci))
        for ci in [5, 9, 12]:
            fmt_pct(ws3.cell(row=r, column=ci))

        if is_total:
            style_total_row(ws3, r, len(headers))
            if acct == "Total Operating Income" or acct == "Total Income":
                income_total_row = r
            if acct == "Total Operating Expense" or acct == "Total Expense":
                expense_total_row = r

        # Track expense line items for charts (non-total, in Expense section)
        if section == "Expense" and not is_total and not acct.startswith("Total"):
            expense_chart_rows.append((r, acct))

        r += 1

    # Column widths
    set_col_widths(ws3, {
        1: 38, 2: 18, 3: 18, 4: 16, 5: 12,
        6: 18, 7: 18, 8: 16, 9: 12, 10: 20, 11: 18, 12: 14,
    })

    # ========== Sheet 4: Charts ==========
    ws4 = out.create_sheet("Charts")
    ws4.sheet_properties.tabColor = "1F4E79"

    # We need to build chart-specific data tables in the Charts sheet
    # because openpyxl charts reference cells.

    # --- Chart 1: Top 10 Expense Categories (2025 Actual vs 2026 Annualized) ---
    # Sort expense rows by 2025 actual (descending), take top 10
    expense_data = []
    for row_num, acct_name in expense_chart_rows:
        d25 = map_2025.get(acct_name)
        d26 = map_2026.get(acct_name)
        ytd_25 = d25["ytd_actual"] if d25 else 0.0
        ytd_26 = d26["ytd_actual"] if d26 else 0.0
        ann_26 = ytd_26 * 12 / MONTHS_2026 if ytd_26 != 0 else 0.0
        # Use max of 2025 or annualized 2026 for sorting (catch items that grew)
        expense_data.append((acct_name, ytd_25, ann_26, max(abs(ytd_25), abs(ann_26))))
    expense_data.sort(key=lambda x: x[3], reverse=True)
    top10 = expense_data[:10]

    # Write data table for Chart 1 starting at row 1
    ws4.cell(row=1, column=1, value="Top 10 Expense Categories")
    ws4.cell(row=1, column=1).font = Font(bold=True, size=12, color="1F4E79")
    ws4.cell(row=2, column=1, value="Account")
    ws4.cell(row=2, column=2, value="2025 Actual")
    ws4.cell(row=2, column=3, value="2026 Annualized")
    style_header(ws4, 2, 3)
    for i, (name, val25, val26, _) in enumerate(top10, 3):
        ws4.cell(row=i, column=1, value=name)
        c2 = ws4.cell(row=i, column=2, value=val25)
        c3 = ws4.cell(row=i, column=3, value=val26)
        fmt_acct(c2)
        fmt_acct(c3)
    set_col_widths(ws4, {1: 35, 2: 18, 3: 20})

    chart1 = BarChart()
    chart1.type = "col"
    chart1.style = 10
    chart1.title = "Top 10 Expense Categories: 2025 Actual vs 2026 Annualized"
    chart1.y_axis.title = "Amount ($)"
    chart1.y_axis.numFmt = '#,##0'
    chart1.x_axis.title = "Account"
    cats = Reference(ws4, min_col=1, min_row=3, max_row=2 + len(top10))
    d1 = Reference(ws4, min_col=2, min_row=2, max_row=2 + len(top10))
    d2 = Reference(ws4, min_col=3, min_row=2, max_row=2 + len(top10))
    chart1.add_data(d1, titles_from_data=True)
    chart1.add_data(d2, titles_from_data=True)
    chart1.set_categories(cats)
    chart1.shape = 4
    chart1.width = 28
    chart1.height = 16
    # Colors
    chart1.series[0].graphicalProperties.solidFill = "1F4E79"
    chart1.series[1].graphicalProperties.solidFill = "E07C24"
    ws4.add_chart(chart1, "E2")

    # --- Chart 2: Income vs Expenses Summary (2025 vs 2026) ---
    # Use Total Income and Total Expense from each year
    inc_25 = map_2025.get("Total Income", {}).get("ytd_actual", 0) or map_2025.get("Total Operating Income", {}).get("ytd_actual", 0)
    exp_25 = map_2025.get("Total Expense", {}).get("ytd_actual", 0) or map_2025.get("Total Operating Expense", {}).get("ytd_actual", 0)
    inc_26_ytd = map_2026.get("Total Income", {}).get("ytd_actual", 0) or map_2026.get("Total Operating Income", {}).get("ytd_actual", 0)
    exp_26_ytd = map_2026.get("Total Expense", {}).get("ytd_actual", 0) or map_2026.get("Total Operating Expense", {}).get("ytd_actual", 0)
    inc_26_ann = inc_26_ytd * 12 / MONTHS_2026 if inc_26_ytd else 0
    exp_26_ann = exp_26_ytd * 12 / MONTHS_2026 if exp_26_ytd else 0

    chart2_start = 15
    ws4.cell(row=chart2_start, column=1, value="Income vs Expenses Summary")
    ws4.cell(row=chart2_start, column=1).font = Font(bold=True, size=12, color="1F4E79")
    ws4.cell(row=chart2_start + 1, column=1, value="Category")
    ws4.cell(row=chart2_start + 1, column=2, value="2025 Actual")
    ws4.cell(row=chart2_start + 1, column=3, value="2026 Annualized")
    style_header(ws4, chart2_start + 1, 3)
    ws4.cell(row=chart2_start + 2, column=1, value="Total Income")
    fmt_acct(ws4.cell(row=chart2_start + 2, column=2, value=inc_25))
    fmt_acct(ws4.cell(row=chart2_start + 2, column=3, value=inc_26_ann))
    ws4.cell(row=chart2_start + 3, column=1, value="Total Expense")
    fmt_acct(ws4.cell(row=chart2_start + 3, column=2, value=exp_25))
    fmt_acct(ws4.cell(row=chart2_start + 3, column=3, value=exp_26_ann))
    ws4.cell(row=chart2_start + 4, column=1, value="Net Income")
    fmt_acct(ws4.cell(row=chart2_start + 4, column=2, value=inc_25 - exp_25))
    fmt_acct(ws4.cell(row=chart2_start + 4, column=3, value=inc_26_ann - exp_26_ann))

    chart2 = BarChart()
    chart2.type = "col"
    chart2.style = 10
    chart2.title = "Income vs Expenses: 2025 Actual vs 2026 Annualized Projection"
    chart2.y_axis.title = "Amount ($)"
    chart2.y_axis.numFmt = '#,##0'
    cats2 = Reference(ws4, min_col=1, min_row=chart2_start + 2, max_row=chart2_start + 4)
    d2a = Reference(ws4, min_col=2, min_row=chart2_start + 1, max_row=chart2_start + 4)
    d2b = Reference(ws4, min_col=3, min_row=chart2_start + 1, max_row=chart2_start + 4)
    chart2.add_data(d2a, titles_from_data=True)
    chart2.add_data(d2b, titles_from_data=True)
    chart2.set_categories(cats2)
    chart2.width = 22
    chart2.height = 14
    chart2.series[0].graphicalProperties.solidFill = "1F4E79"
    chart2.series[1].graphicalProperties.solidFill = "E07C24"
    ws4.add_chart(chart2, "E19")

    # --- Chart 3: Budget Variance — Top Over-Budget Items in 2026 ---
    # Over budget = YTD actual > prorated YTD budget (positive variance for expenses = bad)
    overbudget = []
    for row_num, acct_name in expense_chart_rows:
        d26 = map_2026.get(acct_name)
        if not d26:
            continue
        ytd_26 = d26["ytd_actual"] if d26["ytd_actual"] else 0.0
        bud_26 = d26["annual_budget"] if d26["annual_budget"] else 0.0
        prorated_bud = bud_26 * MONTHS_2026 / 12
        variance = ytd_26 - prorated_bud  # Positive = over budget for expenses
        if variance > 0 and ytd_26 > 0:
            overbudget.append((acct_name, ytd_26, prorated_bud, variance))
    overbudget.sort(key=lambda x: x[3], reverse=True)
    top_over = overbudget[:10]

    chart3_start = 22
    ws4.cell(row=chart3_start, column=1, value="2026 Over-Budget Items (YTD)")
    ws4.cell(row=chart3_start, column=1).font = Font(bold=True, size=12, color="1F4E79")
    ws4.cell(row=chart3_start + 1, column=1, value="Account")
    ws4.cell(row=chart3_start + 1, column=2, value="2026 YTD Actual")
    ws4.cell(row=chart3_start + 1, column=3, value="2026 Prorated Budget")
    ws4.cell(row=chart3_start + 1, column=4, value="Over Budget ($)")
    style_header(ws4, chart3_start + 1, 4)
    set_col_widths(ws4, {4: 18})
    for i, (name, actual, pror_bud, var) in enumerate(top_over, chart3_start + 2):
        ws4.cell(row=i, column=1, value=name)
        fmt_acct(ws4.cell(row=i, column=2, value=actual))
        fmt_acct(ws4.cell(row=i, column=3, value=pror_bud))
        fmt_acct(ws4.cell(row=i, column=4, value=var))

    chart3 = BarChart()
    chart3.type = "col"
    chart3.style = 10
    chart3.title = "2026 Budget Variance: Top Over-Budget Expense Items (YTD)"
    chart3.y_axis.title = "Amount ($)"
    chart3.y_axis.numFmt = '#,##0'
    chart3.x_axis.title = "Account"
    last_over_row = chart3_start + 1 + len(top_over)
    cats3 = Reference(ws4, min_col=1, min_row=chart3_start + 2, max_row=last_over_row)
    d3a = Reference(ws4, min_col=2, min_row=chart3_start + 1, max_row=last_over_row)
    d3b = Reference(ws4, min_col=3, min_row=chart3_start + 1, max_row=last_over_row)
    d3c = Reference(ws4, min_col=4, min_row=chart3_start + 1, max_row=last_over_row)
    chart3.add_data(d3a, titles_from_data=True)
    chart3.add_data(d3b, titles_from_data=True)
    chart3.add_data(d3c, titles_from_data=True)
    chart3.set_categories(cats3)
    chart3.width = 28
    chart3.height = 16
    chart3.series[0].graphicalProperties.solidFill = "C0392B"  # Red - actual
    chart3.series[1].graphicalProperties.solidFill = "2ECC71"  # Green - budget
    chart3.series[2].graphicalProperties.solidFill = "E67E22"  # Orange - over-budget amount
    ws4.add_chart(chart3, "E36")

    # ========== Save ==========
    out.save(OUTPUT)
    print(f"Workbook saved to: {OUTPUT}")
    print(f"  Sheet 1: '2025 Actuals' — full copy of 2025 AppFolio export")
    print(f"  Sheet 2: '2026 YTD' — full copy of 2026 AppFolio export")
    print(f"  Sheet 3: 'Year-over-Year Comparison' — {len(all_accounts)} accounts compared")
    print(f"  Sheet 4: 'Charts' — 3 bar charts with supporting data tables")
    print()
    print("Key findings:")

    # Print some key stats
    net_25 = map_2025.get("Net Income", {}).get("ytd_actual", 0)
    net_26_ytd = map_2026.get("Net Income", {}).get("ytd_actual", 0)
    net_26_ann = net_26_ytd * 12 / MONTHS_2026 if net_26_ytd else 0
    print(f"  2025 Net Income: ${net_25:,.2f}")
    print(f"  2026 Net Income (YTD): ${net_26_ytd:,.2f}")
    print(f"  2026 Net Income (Annualized): ${net_26_ann:,.2f}")
    print()
    if top_over:
        print(f"  Top over-budget items in 2026:")
        for name, actual, pror_bud, var in top_over[:5]:
            print(f"    {name}: ${actual:,.2f} actual vs ${pror_bud:,.2f} budget (${var:,.2f} over)")


if __name__ == "__main__":
    main()
